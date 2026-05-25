"""Leitura estruturada da planilha de entrada."""

from __future__ import annotations

from hashlib import sha256
from pathlib import Path

from openpyxl import load_workbook

from .exceptions import InputDataError
from .models import WorkbookReadResult, WorkbookRow


def read_workbook(workbook_path: str | Path, sheet_name: str, header_row: int, logger=None) -> WorkbookReadResult:
    source_path = Path(workbook_path)
    if not source_path.exists():
        raise InputDataError(f"Workbook not found: {source_path}")

    source_hash = sha256(source_path.read_bytes()).hexdigest()

    workbook = load_workbook(source_path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise InputDataError(f"Worksheet '{sheet_name}' not found in {source_path}")

        worksheet = workbook[sheet_name]
        headers = _read_headers(worksheet, header_row)
        rows = _read_rows(worksheet, headers, header_row)

        if logger is not None:
            logger.info(
                "Workbook loaded: %s | sheet=%s | headers=%d | rows=%d",
                source_path,
                sheet_name,
                len(headers),
                len(rows),
            )

        return WorkbookReadResult(
            source_path=source_path,
            sheet_name=sheet_name,
            headers=headers,
            rows=rows,
            source_hash=source_hash,
        )
    finally:
        workbook.close()


def _read_headers(worksheet, header_row: int) -> list[str]:
    headers = [
        _normalize_cell_value(worksheet.cell(row=header_row, column=column_index).value)
        for column_index in range(1, worksheet.max_column + 1)
    ]

    if not any(headers):
        raise InputDataError("Header row is empty")

    return headers


def _read_rows(worksheet, headers: list[str], header_row: int) -> list[WorkbookRow]:
    rows: list[WorkbookRow] = []

    for row_number in range(header_row + 1, worksheet.max_row + 1):
        values = [
            _normalize_cell_value(worksheet.cell(row=row_number, column=column_index).value)
            for column_index in range(1, len(headers) + 1)
        ]

        if not any(values):
            continue

        row_data = dict(zip(headers, values))
        rows.append(
            WorkbookRow(
                row_number=row_number,
                repository_id=row_data.get("Repositório", ""),
                formal_title=row_data.get("Título formal", ""),
                github_url=row_data.get("GitHub", ""),
                pages_url=row_data.get("Pages", ""),
            )
        )

    return rows


def _normalize_cell_value(value) -> str:
    if value is None:
        return ""
    return str(value).strip()
