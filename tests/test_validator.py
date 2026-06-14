from __future__ import annotations

from openpyxl import load_workbook

from src.config_loader import load_config
from src.data_reader import read_workbook
from src.validator import validate_inputs


def test_validate_inputs_accepts_clean_workbook(temp_workspace):
    config = load_config(temp_workspace["config_path"])
    workbook = read_workbook(
        temp_workspace["input_workbook"],
        config["workbook"]["sheet_name"],
        config["workbook"]["header_row"],
    )

    report = validate_inputs(config, workbook)

    assert report.is_valid is True
    assert report.error_count == 0
    assert report.row_count == 10


def test_validate_inputs_rejects_invalid_github_url(temp_workspace):
    workbook_path = temp_workspace["input_workbook"]
    workbook = load_workbook(workbook_path)
    worksheet = workbook["Repositorios"]
    worksheet[2][2].value = "http://github.com/SILVAThiagoFerreira/enaex-plano-de-voo"
    workbook.save(workbook_path)
    workbook.close()

    config = load_config(temp_workspace["config_path"])
    read_result = read_workbook(
        temp_workspace["input_workbook"],
        config["workbook"]["sheet_name"],
        config["workbook"]["header_row"],
    )
    report = validate_inputs(config, read_result)

    assert report.is_valid is False
    assert any(issue.code == "invalid_github_url_scheme" for issue in report.issues)


def test_validate_inputs_accepts_case_preserved_pages_url(temp_workspace):
    workbook_path = temp_workspace["input_workbook"]
    workbook = load_workbook(workbook_path)
    worksheet = workbook["Repositorios"]
    worksheet["D8"].value = "https://silvathiagoferreira.github.io/ANALISADOR-DE-SISMOGRAMA/"
    workbook.save(workbook_path)
    workbook.close()

    config = load_config(temp_workspace["config_path"])
    read_result = read_workbook(
        temp_workspace["input_workbook"],
        config["workbook"]["sheet_name"],
        config["workbook"]["header_row"],
    )
    report = validate_inputs(config, read_result)

    assert report.is_valid is True
